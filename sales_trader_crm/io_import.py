"""Leitura de planilhas (.csv ou .xlsx) para os imports de clientes/operações/carteira."""
import csv
import io

from openpyxl import load_workbook


def read_tabular_upload(file_storage):
    """file_storage: objeto do Flask (request.files['...']). Retorna (headers, rows)
    onde rows é uma lista de dicts {header: valor}. Suporta .csv e .xlsx."""
    filename = (file_storage.filename or "").lower()
    raw = file_storage.read()

    if filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(raw), data_only=True)
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            return [], []
        rows = []
        for values in rows_iter:
            if values is None or all(v is None for v in values):
                continue
            row = {}
            for h, v in zip(headers, values):
                if not h:
                    continue
                row[h] = "" if v is None else v
            rows.append(row)
        return headers, rows

    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = [dict(r) for r in reader]
    return headers, rows


def guess_column(headers, patterns):
    for h in headers:
        hl = h.strip().lower()
        for pattern in patterns:
            if pattern in hl:
                return h
    return ""


def text_of(row, col):
    if not col:
        return ""
    val = row.get(col)
    return "" if val is None else str(val).strip()


def num_of(row, col):
    if not col:
        return None
    val = row.get(col)
    if val is None or str(val).strip() == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", ".").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None
